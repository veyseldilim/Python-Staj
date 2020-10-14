#Kullanilan kutuphaneler
from openpyxl import Workbook,load_workbook # Excel dosyasi okuma ve yazma islemleri icin
import zeep  # WCF servis
from scipy.spatial import distance #Oklid uzakligi hesaplama

"""
#Is ilani dosyasi okuma
path='c:/Users/veyse/Desktop/pythonexcel/110K_IlanDatasi_23012020_Canli.xlsx' 
wb = load_workbook(path)
ws = wb.active

"""
"""
#CV dosyasi okuma
path4 = 'c:/Users/veyse/Desktop/pythonexcel/data_9EylulCvData_10000.xlsx' 
wb4 = load_workbook(path4)
ws4 = wb4.active

#CV dosyasindan istenilen satirlari okuma ve yeni excel dosyasina yazma
wb5 = Workbook() 
ws5 = wb5.active

#Is dosyasi ilanindan istenilen satirlari okuma ve yeni excel dosyasina yazma
wb6 = Workbook() 
wb6 = wb6.active

"""
#Yeni olusturulan CV dosyasini okuma
path2='c:/Users/veyse/Desktop/pythonexcel/Bilgisayar_Yazilim_CV_Ilanlari.xlsx'
wb2 = load_workbook(path2)
ws2 = wb2.active

#Yeni olusturulan is ilani dosyasini okuma
path3 = 'c:/Users/veyse/Desktop/pythonexcel/Bilgisayar_Muhendisi_Is_Ilanlari.xlsx'
wb3 = load_workbook(path3)
ws3 = wb3.active

"""

#Is ilanlari
#dosyasindan istenilen satirlarin ayiklanmasi ve yeni excel tablosuna yazilmasi
count = 1
flag = False
for i in range(1,ws.max_row+1):
    if str(ws.cell(i,7).value) == 'Bilgisayar Mühendisi':
            for j in range(1,ws.max_column+1):
                c = ws.cell(i,j)
                ws6.cell(count,j).value = c.value
                flag = True
    if flag == True:
        count += 1            
        flag = False

#Yeni olusturulan is ilanlari dosyasini kaydetme
wb6.save('c:/Users/veyse/Desktop/pythonexcel/Bilgisayar_Muhendisi_110K_CV_Ilanlari.xlsx')


"""
"""
#CV dosyasindan istenilen satirlarin ayiklanmasi ve yeni excel tablosuna yazilmasi

count = 1
flag = False
for i in range(1,ws4.max_row+1):
    if str(ws4.cell(i,11).value) == 'Bilgisayar Mühendisliği' or str(ws4.cell(i,11).value) == 'Yazılım Mühendisliği' :
            for j in range(1,ws4.max_column+1):
                c = ws4.cell(i,j)
                ws5.cell(count,j).value = c.value
                flag = True
    if flag == True:
        count += 1            
        flag = False
#Yeni olusturulan CV dosyasini kaydetme
wb5.save('c:/Users/veyse/Desktop/pythonexcel/Bilgisayar_Yazilim_CV_Ilanlari.xlsx')
wb4.close()
wb5.close()
"""
#WCF servisi kullanmak icin link ve client atamasi
print("Connecting to service")
wsdl = "http://193.140.150.95/KariyerServisleri/Service1.svc?singleWsdl"
client = zeep.Client(wsdl=wsdl)

#Olusturulmus CV dosyasindan secilen satirin okunarak attributelerine parcalanmasi
CV_rowNum = 6
cvAttributes = ""
CV = []
# 6 7 8 9 29 32 33 38 39 42 43 49 52 53 58 62 66 70 71 72 73 74 75 76 77 78 79
cells = [6,7,8,9,11,29,32,33,38,39,42,43,49,52,53,58,62,66,70,71,72,73,74,75,76,77,78,79]

for j in range(1,ws2.max_column+1):
    if j in cells:           
        if str(ws2.cell(CV_rowNum,j).value) == 'İstanbul(Asya)' or str(ws2.cell(CV_rowNum,j).value) =='İstanbul(Avr.)':
            sehir = "İstanbul"              
            cvAttributes = cvAttributes + " " + str(sehir)                 
        else:    
            cvAttributes = cvAttributes + " " + str(ws2.cell(CV_rowNum,j).value)  
                   
CV.append(client.service.getRoots(cvAttributes))  
wb2.close()

#Olusturulmus is ilani dosyasindan satirlarin okunarak attributelerine parcalanmasi
ilanlar = []
ilan_cells = [5,6,8,9]
for i in range(1,ws3.max_row+1):
    ilan_explanation =""
    for j in range(3,ws3.max_column+1):
        if j in ilan_cells:
            if str(ws3.cell(i,j).value) == 'İstanbul(Asya)' or str(ws3.cell(i,j).value) =='İstanbul(Avr.)':
                sehir = "İstanbul"
                ilan_explanation += " " + "İstanbul"
            else:
                ilan_explanation += " "+ str(ws3.cell(i,j).value)
    
    ilanlar.append(client.service.getRoots(ilan_explanation))
            
wb3.close()

#Olusturulmus is ilani ozelliklerinin CV ilanindaki karsiliklarinin bulunmasi
dataset = [[] for _ in range(len(ilanlar))]
for x in range(len(ilanlar)):
    for y in range(len(CV[0])):
        if CV[0][y] in ilanlar[x]:
            dataset[x].insert(y,1)
        else:
            dataset[x].insert(y,0)
#CV ilanini da matrix haline getirme
X = []
for x in range(len(CV[0])):
    X.append(1)

#Oklid uzakligi hesaplama
siralanmisİlanlar = [[] for _ in range(len(dataset))]
for i in range(len(dataset)):
    dst = distance.euclidean(X,dataset[i])
    siralanmisİlanlar[i].insert(0,i)
    siralanmisİlanlar[i].insert(1,dst)

#Uzunluga gore siralama
siralanmisİlanlar = sorted(siralanmisİlanlar , key = lambda x: x[1])

#CV dosyasina en yakin ilk 5 is ilaninin ekrana yazdirilmasi
for i in range(5):
    row = siralanmisİlanlar[i][0]
    print(ilanlar[row])

